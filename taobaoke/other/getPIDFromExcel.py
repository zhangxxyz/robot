import openpyxl
import threading
import gevent
from ZK_Model.ZKOrderDataModel import PIDList
import ZK_Model.ZKOrderDataModel

pids = openpyxl.load_workbook("pids.xlsx")
data = pids.get_sheet_by_name('Sheet1')


# for i in range(1,data.max_row):
#     pidName = data.cell(row=i,column=1).value
#     pid = data.cell(row=i,column=2).value
#     pidArray = str(pid).split('_')
#     pids = PIDList()
#     pids.pid = pid
#     pids.pidName = pidName
#     pids.site_id = pidArray[-2]
#     pids.adzone_id = pidArray[-1]
#     pids.saveData(pids)

# re = ZK_Model.ZKOrderDataModel.select([ZK_Model.ZKOrderDataModel.PIDList])
# result = ZK_Model.ZKOrderDataModel.conn.execute(re).first()
# print(result.pid,result.WXID)
